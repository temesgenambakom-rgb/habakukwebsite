import os
import pg8000.native
from urllib.parse import urlparse
from flask import Flask, render_template, request, redirect, session, send_file, g
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

app = Flask(__name__)

# XIYYEEFFANNAA: kun testing/demo qofaaf. Yeroo production irratti fayyadamtan,
# secret_key kana jijjiiruu fi environment variable keessatti kaa'uu qabdu.
app.secret_key = os.environ.get("SECRET_KEY", "habakuk123")

# XIYYEEFFANNAA: username/password kun default qofa. Production irratti
# environment variable fayyadamuu wayya.
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "1234")

# -----------------------------------------------------------------------------
# Database (PostgreSQL - fkn Supabase) - daataan barattootaa fi password-oota
# manneen barnootaa DATABASE_URL keessatti kuufama. Server restart/deploy
# ta'ellee daataan hin badu, sababiin isaas database-n server-icha alatti
# (kaan - Supabase) jiraata.
#
# DATABASE_URL environment variable keessatti kaa'uu qabda, fkn:
# postgresql://user:password@host:5432/dbname
# -----------------------------------------------------------------------------
DATABASE_URL = os.environ.get("DATABASE_URL", "")


def _connect():
    """DATABASE_URL (postgresql://user:pass@host:port/dbname) qooduun
    pg8000.native.Connection tokko banuu."""
    parsed = urlparse(DATABASE_URL)
    return pg8000.native.Connection(
        user=parsed.username,
        password=parsed.password,
        host=parsed.hostname,
        port=parsed.port or 5432,
        database=parsed.path.lstrip("/"),
    )


class DBWrapper:
    """sqlite3.connection fakkeessee db.execute(...).fetchone()/fetchall()
    akka fayyadamnu kan dandeessisu wrapper - kunis code kaan (routes) hunda
    akka jijjiiruu hin barbaachifne godha."""

    def __init__(self, conn):
        self.conn = conn
        self._last_rows = None
        self._last_cols = None

    def _to_named(self, query, params):
        """"?" placeholder-oota gara :p0, :p1 ... (pg8000 named-param
        syntax) jijjiiruu."""
        kwargs = {}
        for i, p in enumerate(params):
            key = f"p{i}"
            query = query.replace("?", f":{key}", 1)
            kwargs[key] = p
        return query, kwargs

    def execute(self, query, params=()):
        query, kwargs = self._to_named(query, params)
        self._last_rows = self.conn.run(query, **kwargs)
        self._last_cols = (
            [c["name"] for c in self.conn.columns] if self.conn.columns else []
        )
        return self

    def executemany(self, query, seq_of_params):
        for params in seq_of_params:
            self.execute(query, params)

    def fetchone(self):
        if not self._last_rows:
            return None
        return dict(zip(self._last_cols, self._last_rows[0]))

    def fetchall(self):
        if not self._last_rows:
            return []
        return [dict(zip(self._last_cols, row)) for row in self._last_rows]

    def commit(self):
        self.conn.commit()


def get_db():
    """Request tokkotti connection database tokko qofa banuu/fayyadamuu."""
    if "db" not in g:
        raw_conn = _connect()
        g.db = DBWrapper(raw_conn)
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.conn.close()


# Manneen barnootaa afran kunneen bulchitan (tartiiba kun /login/school/<id> keessatti fayyadama)
SCHOOLS = [
    "Akuku Secondary School",
    "Gatama Olika Secondary School",
    "Gudina Walal Secondary School",
    "Tajo Secondary School",
]

# Password/username jalqabaa (default) mana barnootaa tokkoon tokkoof - admin booda jijjiiruu danda'a
DEFAULT_SCHOOL_ACCOUNTS = {
    "Akuku Secondary School": ("akuku", "akuku@2024"),
    "Gatama Olika Secondary School": ("gatama", "gatama@2024"),
    "Gudina Walal Secondary School": ("gudina", "gudina@2024"),
    "Tajo Secondary School": ("tajo", "tajo@2024"),
}


def init_db():
    """Table-oota yoo hin jiraatin uumuu (yeroo app jalqabu tokko qofa)."""
    conn = _connect()
    conn.run(
        """
        CREATE TABLE IF NOT EXISTS students (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            gender TEXT NOT NULL,
            grade TEXT NOT NULL,
            department TEXT NOT NULL,
            school TEXT NOT NULL,
            guardian TEXT NOT NULL,
            phone TEXT NOT NULL
        )
        """
    )
    conn.run(
        """
        CREATE TABLE IF NOT EXISTS school_accounts (
            school TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            password TEXT NOT NULL
        )
        """
    )
    conn.commit()

    # Mana barnootaa tokkoon tokkoof account jalqabaa (yoo hin jiraatin qofa) galchuu
    for school, (username, password) in DEFAULT_SCHOOL_ACCOUNTS.items():
        conn.run(
            "INSERT INTO school_accounts (school, username, password) VALUES "
            "(:school, :username, :password) ON CONFLICT (school) DO NOTHING",
            school=school,
            username=username,
            password=password,
        )
    conn.commit()
    conn.close()


# -----------------------------------------------------------------------------
# Gargaarsa (helpers) - eenyummaa fi mirga ilaaluuf
# -----------------------------------------------------------------------------

def is_logged_in():
    return "admin" in session or "school" in session


def is_admin():
    return bool(session.get("admin"))


def scope_school():
    """Admin hunda ni argata (None deebisa = daangaa hin qabu).
    Mana barnootaa tokko yoo seenan, maqaa mana barnootaa isaanii deebisa."""
    if is_admin():
        return None
    return session.get("school")


@app.route("/")
def home():
    return render_template("index.html", schools=SCHOOLS)


# Admin Login (jalqabaa)
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session.clear()
            session["admin"] = True
            return redirect("/dashboard")
        else:
            return render_template("login.html", error="Username ykn Password sirrii miti")

    return render_template("login.html")


# Mana barnootaa filachuun booda seensa (login) mana barnootaa sanaa
@app.route("/login/school/<int:school_id>", methods=["GET", "POST"])
def school_login(school_id):
    if school_id < 0 or school_id >= len(SCHOOLS):
        return redirect("/")

    school_name = SCHOOLS[school_id]

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        db = get_db()
        account = db.execute(
            "SELECT * FROM school_accounts WHERE school = ?", (school_name,)
        ).fetchone()

        if account and username == account["username"] and password == account["password"]:
            session.clear()
            session["school"] = school_name
            return redirect("/dashboard")
        else:
            return render_template(
                "school_login.html",
                school=school_name,
                school_id=school_id,
                error="Username ykn Password sirrii miti",
            )

    return render_template("school_login.html", school=school_name, school_id=school_id)


@app.route("/dashboard")
def dashboard():
    if not is_logged_in():
        return redirect("/")
    return render_template("dashboard.html", is_admin=is_admin(), school_name=session.get("school"))


# Galmeessa barataa
@app.route("/register", methods=["GET", "POST"])
def register():
    if not is_logged_in():
        return redirect("/")

    locked_school = scope_school()

    if request.method == "POST":
        db = get_db()

        # Mana barnootaa tokko yoo ta'e, school-ni isaan galmeessan ofuma isaaniitiin
        # kan cufame (locked) dha - waan biraa filachuu hin danda'an.
        school_value = locked_school if locked_school else request.form["school"]

        db.execute(
            """
            INSERT INTO students (name, gender, grade, department, school, guardian, phone)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                request.form["name"],
                request.form["gender"],
                request.form["grade"],
                request.form["department"],
                school_value,
                request.form["guardian"],
                request.form["phone"],
            ),
        )
        db.commit()

        return redirect("/students")

    return render_template("register.html", schools=SCHOOLS, locked_school=locked_school)


# Kanneen galmaa'an ilaalu (pagination - 20/fuula tokko)
STUDENTS_PER_PAGE = 20


@app.route("/students")
def students():
    if not is_logged_in():
        return redirect("/")

    db = get_db()
    locked_school = scope_school()

    page = request.args.get("page", 1, type=int)
    if page < 1:
        page = 1

    query = request.args.get("q", "").strip()

    # Mana barnootaa tokko yoo ta'e, daangaa isaaniitiin ala hin ilaalan -
    # school_filter isaanii ofumaan mana barnootaa isaaniiti (URL-iin jijjiiruu hin danda'an).
    if locked_school:
        school_filter = locked_school
    else:
        school_filter = request.args.get("school", "").strip()

    where_clauses = []
    params = []

    if school_filter:
        where_clauses.append("school = ?")
        params.append(school_filter)

    if query:
        like_query = f"%{query}%"
        where_clauses.append(
            "(name LIKE ? OR grade LIKE ? OR department LIKE ? OR phone LIKE ?)"
        )
        params.extend([like_query, like_query, like_query, like_query])

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    total_students = db.execute(
        f"SELECT COUNT(*) AS cnt FROM students {where_sql}", params
    ).fetchone()["cnt"]

    total_pages = max(1, (total_students + STUDENTS_PER_PAGE - 1) // STUDENTS_PER_PAGE)

    if page > total_pages:
        page = total_pages

    offset = (page - 1) * STUDENTS_PER_PAGE

    students_page = db.execute(
        f"SELECT * FROM students {where_sql} ORDER BY id LIMIT ? OFFSET ?",
        params + [STUDENTS_PER_PAGE, offset],
    ).fetchall()

    return render_template(
        "students.html",
        students=students_page,
        page=page,
        total_pages=total_pages,
        total_students=total_students,
        query=query,
        school_filter=school_filter,
        schools=SCHOOLS,
        locked_school=locked_school,
    )


# Excel buusuu
@app.route("/export")
def export():
    if not is_logged_in():
        return redirect("/")

    db = get_db()
    locked_school = scope_school()

    if locked_school:
        all_students = db.execute(
            "SELECT * FROM students WHERE school = ? ORDER BY id", (locked_school,)
        ).fetchall()
    else:
        all_students = db.execute("SELECT * FROM students ORDER BY id").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Barattoota"

    # Mata duree (header)
    headers = ["ID", "Maqaa", "Koorniyaa", "Kutaa", "Department", "Mana Barnootaa", "Wabii", "Bilbila"]
    ws.append(headers)

    # Mata duree bareechuu (bold, background green, font white)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Daataa barattootaa
    for student in all_students:
        ws.append([
            student["id"],
            student["name"],
            student["gender"],
            student["grade"],
            student["department"],
            student["school"],
            student["guardian"],
            student["phone"]
        ])

    # Column width-oota daataa isaaniif mijeessuu
    for col_cells in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=10
        )
        column_letter = col_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 4

    # Mata duree cufaa jiru (freeze) akka scroll gootan yeroo hin badne
    ws.freeze_panes = "A2"

    # File memory keessatti uumuu (hard disk irratti hin kuufamu)
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Maqaa file-ii guyyaa har'aa waliin (fkf. barattoota_2026-07-22.xlsx)
    today = datetime.now().strftime("%Y-%m-%d")

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"barattoota_{today}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Filannoowwan sirrii (register.html waliin walsimu - validation-iif)
VALID_GENDERS = {"Dhiira", "Dhalaa"}
VALID_GRADES = {"9", "10", "11", "12"}
VALID_DEPARTMENTS = {"Hin qabu", "Natural", "Social"}

# Import-iif mata duree jajjabeeffaman (ID hin qabatu - automatically uumama)
IMPORT_HEADERS = ["Maqaa", "Koorniyaa", "Kutaa", "Department", "Mana Barnootaa", "Wabii", "Bilbila"]


# Template Excel duwwaa buusuu (barattoota baayina guutuuf)
@app.route("/students/template")
def students_template():
    if not is_logged_in():
        return redirect("/")

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws.append(IMPORT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Fakkeenya (example) tokko akka hordofan gargaaruuf
    ws.append(["Fakkeenya Abdii", "Dhiira", "10", "Natural", SCHOOLS[0], "Obboleessa", "0911223344"])

    for col_cells in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=10
        )
        column_letter = col_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 4

    ws.freeze_panes = "A2"

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="template_barattoota.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Excel guutame upload godhanii barattoota baayinaan galmeessuu
@app.route("/students/import", methods=["GET", "POST"])
def students_import():
    if not is_logged_in():
        return redirect("/")

    locked_school = scope_school()

    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")

        if not uploaded_file or uploaded_file.filename == "":
            return render_template("import.html", error="Maaloo file Excel filadhaa.")

        if not uploaded_file.filename.lower().endswith((".xlsx", ".xlsm")):
            return render_template("import.html", error="File-ni .xlsx qofa fudhatama.")

        try:
            wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
            ws = wb.active
        except Exception:
            return render_template("import.html", error="File-icha banuu hin dandeenye. Excel sirrii ta'e fayyadamaa.")

        db = get_db()
        added_count = 0
        skipped_rows = []
        rows_to_insert = []

        # Row 1 mata duree waan ta'eef, row 2 irraa eegalla
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Row duwwaa yoo ta'e darbi
            if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            gender = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            grade = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            department = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            school = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            guardian = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
            phone = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""

            # Mana barnootaa tokko yoo ta'e, school-ni isaanii ofuma isaaniitiin cufame - column
            # "Mana Barnootaa" excel keessaa dhiisanii mana barnootaa isaanii qofatti galmeessu.
            if locked_school:
                school = locked_school

            # Validation - fields barbaachisoo hunda guutamuu qabu, filannoowwanis sirrii ta'uu qabu
            if not all([name, gender, grade, department, school, guardian, phone]):
                skipped_rows.append(f"Row {row_num}: odeeffannoo hin guutamne")
                continue

            if gender not in VALID_GENDERS:
                skipped_rows.append(f"Row {row_num}: Koorniyaa '{gender}' sirrii miti (Dhiira/Dhalaa qofa)")
                continue

            if grade not in VALID_GRADES:
                skipped_rows.append(f"Row {row_num}: Kutaa '{grade}' sirrii miti (9-12 qofa)")
                continue

            if department not in VALID_DEPARTMENTS:
                skipped_rows.append(f"Row {row_num}: Department '{department}' sirrii miti")
                continue

            if school not in SCHOOLS:
                skipped_rows.append(f"Row {row_num}: Mana Barnootaa '{school}' sirrii miti")
                continue

            rows_to_insert.append((name, gender, grade, department, school, guardian, phone))
            added_count += 1

        if rows_to_insert:
            db.executemany(
                """
                INSERT INTO students (name, gender, grade, department, school, guardian, phone)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                rows_to_insert,
            )
            db.commit()

        return render_template(
            "import.html",
            success_count=added_count,
            skipped_rows=skipped_rows
        )

    return render_template("import.html")


# Barataa ID isaatiin barbaaduu (edit/delete-iif gargaara)
def find_student(student_id):
    db = get_db()
    return db.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()


# Barataa tokko fooyyessuu (Edit)
@app.route("/students/edit/<int:student_id>", methods=["GET", "POST"])
def students_edit(student_id):
    if not is_logged_in():
        return redirect("/")

    locked_school = scope_school()

    student = find_student(student_id)
    if student is None:
        return redirect("/students")

    # Mana barnootaa tokko, barataa mana barnootaa biraa fooyyessuu hin danda'u
    if locked_school and student["school"] != locked_school:
        return redirect("/students")

    if request.method == "POST":
        db = get_db()
        school_value = locked_school if locked_school else request.form["school"]
        db.execute(
            """
            UPDATE students
            SET name = ?, gender = ?, grade = ?, department = ?, school = ?, guardian = ?, phone = ?
            WHERE id = ?
            """,
            (
                request.form["name"],
                request.form["gender"],
                request.form["grade"],
                request.form["department"],
                school_value,
                request.form["guardian"],
                request.form["phone"],
                student_id,
            ),
        )
        db.commit()

        return redirect("/students")

    return render_template("edit.html", student=student, schools=SCHOOLS, locked_school=locked_school)


# Barataa tokko haquu (Delete)
@app.route("/students/delete/<int:student_id>", methods=["POST"])
def students_delete(student_id):
    if not is_logged_in():
        return redirect("/")

    locked_school = scope_school()
    student = find_student(student_id)

    # Mana barnootaa tokko, barataa mana barnootaa biraa haquu hin danda'u
    if locked_school and (student is None or student["school"] != locked_school):
        return redirect("/students")

    db = get_db()
    db.execute("DELETE FROM students WHERE id = ?", (student_id,))
    db.commit()

    return redirect("/students")


# Admin qofti password manneen barnootaa jijjiiruu danda'a
@app.route("/admin/passwords", methods=["GET", "POST"])
def admin_passwords():
    if not is_admin():
        return redirect("/")

    db = get_db()
    message = None

    if request.method == "POST":
        for i, school in enumerate(SCHOOLS):
            new_username = request.form.get(f"username_{i}", "").strip()
            new_password = request.form.get(f"password_{i}", "").strip()
            if new_username and new_password:
                db.execute(
                    "UPDATE school_accounts SET username = ?, password = ? WHERE school = ?",
                    (new_username, new_password, school),
                )
        db.commit()
        message = "Jijjiiramni milkaa'ee kuufame!"

    accounts_rows = db.execute("SELECT * FROM school_accounts").fetchall()
    accounts_by_school = {row["school"]: row for row in accounts_rows}
    ordered_accounts = [accounts_by_school[s] for s in SCHOOLS]

    return render_template("admin_passwords.html", accounts=ordered_accounts, message=message)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# Database jalqaba (table-oota yoo hin jiraatin uumuu) - app import/run yeroo hundaa
init_db()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
